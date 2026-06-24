"""Generates PolicyBot_Sample_QA.xlsx with 3 sheets:
  1. Sample Questions  — 20 rows, Question | Expected_Answer_Summary | Policy_Document | Page_Number
  2. Evaluation Sheet  — 20 rows, Question | AI_Response | Your_Evaluation | Notes
  3. Answer Key        — hidden, Question | Correct_Answer | Known_Flaw_Type | Explanation

Run directly:  python generate_excel.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent / "sample_data"
OUTPUT_FILE = OUTPUT_DIR / "PolicyBot_Sample_QA.xlsx"

QA_DATA = [
    {
        "question": "How many earned leaves can an employee carry forward per year?",
        "expected": "Up to 30 days of EL can be carried forward; balance above 30 lapses on 31 December.",
        "doc": "HR_Leave_Policy.pdf",
        "page": 3,
        "correct": "Up to 30 days of Earned Leave can be carried forward to the next year. Any EL balance beyond 30 days as on 31 December lapses automatically unless encashed.",
        "flaw": "wrong_number",
        "explanation": "A flawed answer may state 25 or 35 instead of 30 days.",
    },
    {
        "question": "What is the per diem allowance for international travel?",
        "expected": "USD 75/day for USA/UK/Western Europe/Singapore/Japan, USD 60/day for other countries.",
        "doc": "HR_Travel_Reimbursement_Policy.pdf",
        "page": 4,
        "correct": "USD 75 per day for USA, UK, Western Europe, Singapore, Japan; USD 60 per day for other countries.",
        "flaw": "wrong_number",
        "explanation": "Flawed answers may shift the figure to USD 65 or USD 80.",
    },
    {
        "question": "How many dependents are covered under medical insurance?",
        "expected": "Maximum 2 dependents under the base GMC policy in addition to the employee.",
        "doc": "HR_Medical_Insurance_Policy.pdf",
        "page": 3,
        "correct": "An employee may enroll a maximum of 2 dependents under the base policy. Eligible dependents are spouse and up to 2 children.",
        "flaw": "missing_clause",
        "explanation": "Flawed answer may omit that parents are not covered under the base policy.",
    },
    {
        "question": "Who approves expense claims above Rs. 50,000?",
        "expected": "The relevant Vice President approves claims above Rs. 50,000; claims above Rs. 2,00,000 also need Finance Controller concurrence.",
        "doc": "Finance_Expense_Reimbursement_Policy.pdf",
        "page": 2,
        "correct": "Claims above Rs. 50,000 require approval from the relevant Vice President. Claims above Rs. 2,00,000 additionally need concurrence from the Finance Controller.",
        "flaw": "wrong_approver",
        "explanation": "Flawed answer may attribute approval to Business Unit Head or HR.",
    },
    {
        "question": "Can an employee on probation avail WFH?",
        "expected": "Probationers are NOT eligible for standard WFH; exceptions need HR Head approval for genuine reasons.",
        "doc": "WFH_Remote_Work_Policy.pdf",
        "page": 2,
        "correct": "Employees on probation are NOT eligible for standard WFH benefit. Exceptions may be granted only with HR Head approval for genuine reasons such as medical advisories.",
        "flaw": "missing_clause",
        "explanation": "Flawed answer may omit the probation exclusion entirely.",
    },
    {
        "question": "What is the maximum number of WFH days per week for confirmed employees?",
        "expected": "Maximum 2 WFH days per week.",
        "doc": "WFH_Remote_Work_Policy.pdf",
        "page": 2,
        "correct": "Confirmed employees may avail a maximum of 2 WFH days per week. The remaining 3 days must be from the office unless an approved exception is in place.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may state 3 days instead of 2.",
    },
    {
        "question": "How is the annual performance rating calculated?",
        "expected": "Weighted: 70% goal achievement and 30% behavioural competencies, mapped to a 5-point scale.",
        "doc": "Performance_Management_Policy.pdf",
        "page": 2,
        "correct": "Final ratings are computed as a weighted score of 70% goal achievement and 30% behavioural competencies, then mapped to the 5-point scale (1-Below to 5-Outstanding).",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may flip the weights or alter them slightly (e.g., 60/40).",
    },
    {
        "question": "What documents are required for medical reimbursement claims?",
        "expected": "Original receipts, discharge summary, and doctor's prescription within 45 days of discharge.",
        "doc": "HR_Medical_Insurance_Policy.pdf",
        "page": 5,
        "correct": "Original receipts, discharge summary, and the doctor's prescription must be submitted within 45 days of discharge for reimbursement claims.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may say 30 or 60 days instead of 45.",
    },
    {
        "question": "What is the policy on software installation on company devices?",
        "expected": "Only software from the IT-approved catalogue; unapproved software needs a Software Request Ticket and IT Security review.",
        "doc": "IT_Acceptable_Use_Policy.pdf",
        "page": 3,
        "correct": "Only software from the IT-approved catalogue may be installed. Self-installation through the catalogue is permitted for whitelisted titles. Anything outside requires a Software Request Ticket and IT Security review.",
        "flaw": "missing_clause",
        "explanation": "Flawed answer may omit the prohibition on pirated/cracked software or personal cloud storage.",
    },
    {
        "question": "How many casual leaves can be clubbed with earned leaves?",
        "expected": "CL cannot be clubbed with EL under any circumstance.",
        "doc": "HR_Leave_Policy.pdf",
        "page": 5,
        "correct": "Casual Leave cannot be clubbed with Earned Leave under any circumstances. It may be clubbed with weekly off or public holidays.",
        "flaw": "wrong_approver",
        "explanation": "Flawed answer may claim clubbing is allowed with manager approval.",
    },
    {
        "question": "What is the sum insured under the base medical policy?",
        "expected": "Rs. 3,00,000 family floater; top-up of Rs. 5,00,000 available at employee-paid premium.",
        "doc": "HR_Medical_Insurance_Policy.pdf",
        "page": 2,
        "correct": "The Sum Insured is Rs. 3,00,000 per family on a floater basis, with an optional top-up of Rs. 5,00,000 at subsidised employee-paid premium during the April enrollment window.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may misstate as Rs. 2,50,000 or Rs. 4,00,000.",
    },
    {
        "question": "What is the timeline for reporting a data breach?",
        "expected": "Within 2 hours to the Incident Response Team; regulatory notification (if applicable) within 72 hours.",
        "doc": "IT_Data_Security_Policy.pdf",
        "page": 4,
        "correct": "Any suspected or confirmed data breach must be reported within 2 hours to the Incident Response Team. Regulatory breach notifications, where applicable, must be initiated within 72 hours.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may shift to 4 hours, 24 hours, or change the 72 hour regulatory window.",
    },
    {
        "question": "What is the gift limit employees may accept from vendors or clients?",
        "expected": "Token gifts up to Rs. 2,500; cash and gift cards are never acceptable.",
        "doc": "HR_Code_of_Conduct.pdf",
        "page": 3,
        "correct": "Employees may accept token gifts from clients and vendors not exceeding Rs. 2,500 in value. Cash, gift cards, and equivalent instruments must NEVER be accepted regardless of value.",
        "flaw": "missing_clause",
        "explanation": "Flawed answer may omit the absolute prohibition on cash and gift cards.",
    },
    {
        "question": "What approvals are required for a Purchase Order of Rs. 6,00,000?",
        "expected": "Procurement Head and Finance Controller approval; PO above Rs. 5,00,000 needs VP approval too.",
        "doc": "Finance_Procurement_Policy.pdf",
        "page": 2,
        "correct": "Purchase Orders above Rs. 5,00,000 require approval from the relevant Vice President. (Procurement Head and Finance Controller approve up to Rs. 5,00,000.) POs above Rs. 25,00,000 additionally require CFO approval.",
        "flaw": "wrong_approver",
        "explanation": "Flawed answer may say Finance Controller alone is sufficient.",
    },
    {
        "question": "How many paternity leave days are employees entitled to?",
        "expected": "10 working days, to be availed within 3 months of birth, covers up to 2 children.",
        "doc": "HR_Leave_Policy.pdf",
        "page": 6,
        "correct": "Male employees are entitled to 10 working days of paid Paternity Leave, to be availed within 3 months of the birth of the child. Covers up to 2 children.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may state 5, 7, or 15 days instead of 10.",
    },
    {
        "question": "What is the duration of the Performance Improvement Plan?",
        "expected": "90 days; may be extended by 30 days if needed.",
        "doc": "Performance_Management_Policy.pdf",
        "page": 3,
        "correct": "Employees rated Below Expectations are placed on a formal PIP of 90 days duration. Unsuccessful completion may lead to extension by 30 days, role change, or separation.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may say 60 days or 120 days.",
    },
    {
        "question": "Is moonlighting allowed at Meridian Technologies?",
        "expected": "External employment requires prior written approval from BU Head and HR; unapproved is a serious breach.",
        "doc": "HR_Code_of_Conduct.pdf",
        "page": 2,
        "correct": "External employment (moonlighting) requires prior written approval from the Business Unit Head and HR. Unapproved external employment is treated as a serious breach of contract.",
        "flaw": "missing_clause",
        "explanation": "Flawed answer may suggest it is freely allowed.",
    },
    {
        "question": "What is the maximum continuous earned leave that can be taken at a stretch?",
        "expected": "15 working days; longer requires Business Unit Head approval.",
        "doc": "HR_Leave_Policy.pdf",
        "page": 3,
        "correct": "Maximum continuous EL is 15 working days; longer durations require Business Unit Head approval.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may state 10 or 20 days.",
    },
    {
        "question": "How many quotations are required for procurement above Rs. 1,00,000?",
        "expected": "Minimum of 3 written quotations from empanelled vendors.",
        "doc": "Finance_Procurement_Policy.pdf",
        "page": 3,
        "correct": "Procurement above Rs. 1,00,000 requires a minimum of 3 written quotations from empanelled vendors.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may say 2 or 4 quotations.",
    },
    {
        "question": "What rating distribution is enforced during calibration?",
        "expected": "Not more than 15% combined Outstanding+Exceeds; not more than 10% combined Below+Partially Meets.",
        "doc": "Performance_Management_Policy.pdf",
        "page": 3,
        "correct": "Calibration guidelines at Business Unit level: not more than 15% combined rated Outstanding or Exceeds, and not more than 10% combined rated Below or Partially Meets.",
        "flaw": "wrong_number",
        "explanation": "Flawed answer may give percentages like 20% / 5% or other shifted distributions.",
    },
]


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
ALT_FILL = PatternFill("solid", fgColor="F0F4FA")


def _write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_data_row(ws, row_idx, ncols, alt):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sample Questions"

    _write_header(ws1, ["Question", "Expected_Answer_Summary", "Policy_Document", "Page_Number"])
    for i, item in enumerate(QA_DATA, start=2):
        ws1.cell(row=i, column=1, value=item["question"])
        ws1.cell(row=i, column=2, value=item["expected"])
        ws1.cell(row=i, column=3, value=item["doc"])
        ws1.cell(row=i, column=4, value=item["page"])
        _style_data_row(ws1, i, 4, alt=(i % 2 == 0))
    _set_col_widths(ws1, [55, 60, 38, 12])

    ws2 = wb.create_sheet("Evaluation Sheet")
    _write_header(ws2, ["Question", "AI_Response", "Your_Evaluation (Correct/Flawed)", "Notes"])
    for i, item in enumerate(QA_DATA, start=2):
        ws2.cell(row=i, column=1, value=item["question"])
        ws2.cell(row=i, column=2, value="")
        ws2.cell(row=i, column=3, value="")
        ws2.cell(row=i, column=4, value="")
        _style_data_row(ws2, i, 4, alt=(i % 2 == 0))
    _set_col_widths(ws2, [55, 60, 28, 40])

    ws3 = wb.create_sheet("Answer Key")
    _write_header(ws3, ["Question", "Correct_Answer", "Known_Flaw_Type", "Explanation"])
    for i, item in enumerate(QA_DATA, start=2):
        ws3.cell(row=i, column=1, value=item["question"])
        ws3.cell(row=i, column=2, value=item["correct"])
        ws3.cell(row=i, column=3, value=item["flaw"])
        ws3.cell(row=i, column=4, value=item["explanation"])
        _style_data_row(ws3, i, 4, alt=(i % 2 == 0))
    _set_col_widths(ws3, [50, 65, 22, 50])
    ws3.sheet_state = "hidden"

    wb.save(OUTPUT_FILE)
    print(f"  generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
